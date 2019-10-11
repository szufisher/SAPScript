import os,win32com.client
import time
from openpyxl import load_workbook
from datetime import datetime
import schedule
from utils import logon_sap,start_sap,close_sap, connect_db, close_db,get_configer,get_logger, timestamp, send_email

cf =get_configer('asset_interface.conf')
logger = get_logger('asset_interface.log')

def execute_transaction(session, tcode, company, sort_variant, conn, cursor):
    def get_workbook(fullname):
        try:
            return load_workbook(fullname)
        except:
            return False
            
    # Download Excel
    session.findById("wnd[0]/tbar[0]/okcd").Text = tcode  #"/nS_ALR_87011990"
    session.findById("wnd[0]").sendVKey(0)
    session.findById("wnd[0]/usr/radXEINZEL").Select()
    session.findById("wnd[0]/usr/ctxtBUKRS-LOW").Text = company #"cn10"
    session.findById("wnd[0]/usr/ctxtBEREICH1").Text = "60"
    session.findById("wnd[0]/usr/ctxtSRTVR").Text = sort_variant #"0002"
    session.findById("wnd[0]/tbar[1]/btn[8]").press()
    session.findById("wnd[0]/mbar/menu[0]/menu[1]/menu[1]").Select()
    #session.findById("wnd[1]/usr/radRB_1").Select()
    session.findById("wnd[1]/usr/radRB_OTHERS").select()
    session.findById("wnd[1]/usr/cmbG_LISTBOX").key = "31"
    session.findById("wnd[1]/tbar[0]/btn[0]").press()
    path = session.findById("wnd[1]/usr/ctxtDY_PATH").Text        
    filename = "S_ALR_87011990_%s.xlsx" %(timestamp())
    session.findById("wnd[1]/usr/ctxtDY_FILENAME").Text = filename
    session.findById("wnd[1]/tbar[0]/btn[0]").press()    
    fullname = os.path.join(path ,filename)    
    
    # upload to db
    i = 30
    while i > 0:
        wbk = get_workbook(fullname)
        if wbk:            
            break
        else:            
            i -= 1            
            time.sleep(1)    
       
    sht= wbk.active  
    last_row = sht.max_row
    total_records_updated = last_row - 1
    col_check_ok = True
    i = 0
    fields = cf.get('transaction','fields')
    fields = fields.split(',')
    for field in fields:
        i += 1
        if field != sht.cell(row=1, column=i).value:
            col_check_ok = False
            print('field:%s <> downloaded field:%s' % (field, sht.cell(row=1, column=i).value))
            break
    if not col_check_ok:
        logger.info('Fields sequence should be same as in asset_interface.conf file, field index %s ' %(i))        
        return
    
    cursor.execute ("delete from app_fd_F01_AssetMaster")
    db_fields =['id',
                'c_AssetNo',
                'c_SubNumber',
                'c_AssetClass',
                'c_AssetDescription',
                'c_CompanyCode',
                'c_Plant',
                ' c_CostCenter',
                'c_AssetOwnerNo',
                'c_CapitalizedDate',
                'c_DeactivationDate',
                'c_Currency',
                'c_CurrBkVal',
                'c_CurrentAPC',
                'c_AccumulDep',
                'dateCreated',
                'createdBy']
    s = "INSERT into app_fd_F01_AssetMaster (%s) VALUES (%s) " %(','.join(db_fields), ','.join(['?']*len(db_fields)))    
    record_value = []
    for j in xrange(2,last_row+1):
        if sht.cell(row=j, column=1).value:
            imod = j % 1000
            record_value.append(get_record_value(sht, j))
        if j == last_row or imod == 0:
            cursor.executemany(s, record_value)
            cursor.commit()
            record_value =[]
    close_db(conn)
    wbk.close()
    
    xl = win32com.client.GetObject(Class='Excel.application')
    xl.quit()
    return total_records_updated
    

def get_record_value(sht, irow):   
    rec = [sht.cell(row=irow, column=1).value]
    for k in xrange(1,15):
        rec.append(sht.cell(row=irow, column=k).value)    
    rec.extend([datetime.now(),os.environ['username']])
    return rec

def job():
    try:
        print('%s started running the job...' % datetime.now())    
        short_cut_file =cf.get('saplogon','short_cut_file')
        popup_win_title=cf.get('saplogon','popup_win_title')
        pin =cf.get('saplogon','pin')
        wait_sec =cf.get('saplogon','wait_sec')
        logon_sap(short_cut_file, popup_win_title, pin, wait_sec)
        time.sleep(2)                
        j = 5
        while j > 0:
            session = start_sap()
            if session:
                break
            else:
                j-= 1
                time.sleep(1)
                
        if session:
            ip = cf.get('db','ip')
            db = cf.get('db','db')
            conn, cursor = connect_db(ip, db)
            
            tcode =cf.get('transaction','tcode')
            company =cf.get('transaction','company')
            sort_variant =cf.get('transaction','sort_variant')
            total_records_updated = execute_transaction(session,tcode,company,sort_variant, conn, cursor)
            
            close_sap(session)
            send_email('%s Asset updated' % total_records_updated, "asset_interface@b.com")
        else:
            send_email("Failed logon SAP", 'asset_interface@b.com')
        print('%s finished running the job...' % datetime.now())                
    except Exception, e:
        send_email("Asset interface run with error %s" % str(e), 'asset_interface@b.com')
        raise
        
def main():
    print('started..')
    run_at=cf.get('schedule','RunAt')
    #mailto=cf.get('mail','MailTo')    
    runat = run_at.split(';')
    for r in runat:
        schedule.every().day.at(r).do(job)
    print('%s waiting for pending job at %s' %(datetime.now(),runat))    
    while True:
        schedule.run_pending()
        time.sleep(1)
        
if __name__ == "__main__":
    main()    
